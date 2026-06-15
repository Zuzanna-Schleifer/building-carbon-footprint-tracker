# ============================================================
# Projekt Data 3 — Carbon Footprint Tracker
# Notebook 04: Eksport raportu CSRD (Excel 4 zakładki)
# Author: Zuzanna Schleifer · github.com/Zuzanna-Schleifer
# ============================================================

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

emissions  = pd.read_parquet('../data/processed/emissions_calculated.parquet')
audit      = pd.read_parquet('../data/processed/audit_trail.parquet')
buildings  = pd.read_parquet('../data/raw/buildings.parquet')
annual_kpi = pd.read_csv('../outputs/annual_kpi.csv')
ef_df      = pd.read_csv('../data/raw/emission_factors.csv')


# ── Dane do zakładek ─────────────────────────────────────────
def build_kpi_summary(annual_kpi, emissions, buildings):
    last = annual_kpi[annual_kpi['year'] == 2023].iloc[0]
    prev = annual_kpi[annual_kpi['year'] == 2022].iloc[0]
    yoy  = (last['total'] - prev['total']) / prev['total'] * 100
    sqm_total = buildings['sqm'].sum()
    intensity = last['total'] * 1000 / sqm_total

    taxonomy_ok = (emissions
        .assign(year=emissions['month'].dt.year).query('year == 2023')
        .groupby('building_id')['kgco2eq'].sum().reset_index()
        .merge(buildings[['building_id', 'sqm']], on='building_id')
        .assign(intensity_kg=lambda x: x['kgco2eq'] / x['sqm'])
        .assign(ok=lambda x: x['intensity_kg'] < 25)
    )['ok'].mean()

    rows = [
        ['ESRS E1-6', 'Łączne emisje Scope 1',
         f"{last.get('1', 0):.1f} tCO₂eq", '2023', 'Licznik gazu + F-gazy'],
        ['ESRS E1-6', 'Łączne emisje Scope 2 (location-based)',
         f"{last.get('2_location', 0):.1f} tCO₂eq", '2023', 'Licznik energii × KOBiZE'],
        ['ESRS E1-6', 'Łączne emisje Scope 2 (market-based)',
         f"{last.get('2_market', 0):.1f} tCO₂eq", '2023', 'Licznik energii × GO/PPA'],
        ['ESRS E1-6', 'Łączne emisje Scope 3',
         f"{last.get('3', 0):.1f} tCO₂eq", '2023', 'Dojazdy + odpady + woda'],
        ['ESRS E1-6', 'Łączne emisje (Scope 1+2+3)',
         f"{last['total']:.1f} tCO₂eq", '2023', 'Suma Scope 1+2+3'],
        ['ESRS E1-4', 'Intensywność emisji',
         f"{intensity:.2f} kgCO₂eq/m²", '2023', 'Łączne emisje / GLA'],
        ['ESRS E1-6', 'Zmiana YoY',
         f"{yoy:+.1f}%", '2022→2023', 'Porównanie do roku poprzedniego'],
        ['EU Taxonomy', 'Budynki zgodne z EU Taxonomy',
         f"{taxonomy_ok:.0%}", '2023', 'Intensity < 25 kgCO₂eq/m²'],
        ['SBTi', 'Cel redukcji emisji', '-4.2% rocznie', 'Baza: 2021',
         'Science Based Targets initiative'],
    ]
    return pd.DataFrame(rows, columns=['Standard', 'KPI', 'Wartość', 'Okres', 'Źródło danych'])


def style_header(ws, row_num, fill_color='1D6FA5'):
    fill = PatternFill('solid', fgColor=fill_color)
    font = Font(bold=True, color='FFFFFF', size=11)
    for cell in ws[row_num]:
        if cell.value:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def style_data_rows(ws, start_row, end_row, col_count):
    fill_odd  = PatternFill('solid', fgColor='F1EFE8')
    fill_even = PatternFill('solid', fgColor='FFFFFF')
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=col_count):
        fill = fill_odd if row[0].row % 2 == 0 else fill_even
        for cell in row:
            cell.fill = fill
            cell.alignment = Alignment(vertical='center')


def set_column_widths(ws, widths):
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def write_df_to_sheet(ws, df):
    for r in dataframe_to_rows(df, index=False, header=False):
        ws.append(r)


# ── Przygotowanie danych ──────────────────────────────────────
kpi_summary = build_kpi_summary(annual_kpi, emissions, buildings)

detail = (emissions
    .assign(year=emissions['month'].dt.year)
    .groupby(['building_id', 'year', 'scope_label', 'source_type'])['kgco2eq'].sum()
    .reset_index()
    .merge(buildings[['building_id', 'primary_use', 'city', 'sqm']], on='building_id')
    .assign(tco2eq=lambda x: (x['kgco2eq'] / 1000).round(3))
    .sort_values(['building_id', 'year', 'scope_label'])
    [['building_id', 'primary_use', 'city', 'sqm', 'year', 'scope_label', 'source_type', 'tco2eq']]
)

audit_export = (audit
    .assign(year=audit['month'].dt.year)
    .groupby(['building_id', 'year', 'scope_label', 'source_type',
              'emission_factor', 'ef_unit', 'ef_source'])
    .agg(total_quantity=('quantity', 'sum'), total_kgco2eq=('kgco2eq', 'sum'),
         calculated_at=('calculated_at', 'first'))
    .reset_index()
    .assign(total_tco2eq=lambda x: (x['total_kgco2eq'] / 1000).round(3))
    .sort_values(['building_id', 'year'])
)


# ── Generowanie Excela ────────────────────────────────────────
def generate_csrd_report(kpi_summary, detail, audit_export, ef_df, output_path):
    wb = openpyxl.Workbook()

    # Zakładka 1: KPI Summary
    ws1 = wb.active
    ws1.title = '1. KPI Summary'
    ws1.row_dimensions[1].height = 40
    ws1['A1'] = 'CSRD Carbon Footprint Report — Portfolio nieruchomości 2023'
    ws1['A1'].font = Font(bold=True, size=13)
    ws1.merge_cells('A1:E1')
    ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws1.append(['Standard', 'KPI', 'Wartość', 'Okres', 'Źródło danych'])
    style_header(ws1, 2)
    write_df_to_sheet(ws1, kpi_summary)
    style_data_rows(ws1, 3, 3 + len(kpi_summary), 5)
    set_column_widths(ws1, [14, 42, 20, 14, 32])
    ws1.freeze_panes = 'A3'

    # Zakładka 2: Dane szczegółowe
    ws2 = wb.create_sheet('2. Dane szczegółowe')
    ws2.append(['Building ID', 'Typ', 'Miasto', 'GLA (m²)',
                'Rok', 'Scope', 'Źródło emisji', 'tCO₂eq'])
    style_header(ws2, 1, fill_color='0F6E56')
    write_df_to_sheet(ws2, detail)
    style_data_rows(ws2, 2, 2 + len(detail), 8)
    set_column_widths(ws2, [12, 14, 12, 10, 6, 14, 22, 10])
    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = ws2.dimensions

    # Zakładka 3: Audit Trail
    ws3 = wb.create_sheet('3. Audit Trail')
    ws3.append(['Building ID', 'Rok', 'Scope', 'Źródło emisji',
                'Emission factor', 'Jednostka EF', 'Źródło EF',
                'Ilość łączna', 'tCO₂eq', 'Data obliczeń'])
    style_header(ws3, 1, fill_color='3C3489')
    audit_cols = ['building_id', 'year', 'scope_label', 'source_type',
                  'emission_factor', 'ef_unit', 'ef_source',
                  'total_quantity', 'total_tco2eq', 'calculated_at']
    write_df_to_sheet(ws3, audit_export[audit_cols])
    style_data_rows(ws3, 2, 2 + len(audit_export), 10)
    set_column_widths(ws3, [12, 6, 14, 22, 14, 18, 26, 14, 10, 14])
    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = ws3.dimensions

    # Zakładka 4: Emission Factors
    ws4 = wb.create_sheet('4. Emission Factors')
    ws4.append(['Źródło emisji', 'Współczynnik', 'Jednostka', 'Baza danych'])
    style_header(ws4, 1, fill_color='854F0B')
    write_df_to_sheet(ws4, ef_df[['source_type', 'factor', 'unit', 'source']])
    style_data_rows(ws4, 2, 2 + len(ef_df), 4)
    set_column_widths(ws4, [26, 14, 22, 28])
    ws4.freeze_panes = 'A2'

    wb.save(output_path)
    print(f"Raport zapisany: {output_path}")


generate_csrd_report(
    kpi_summary, detail, audit_export, ef_df,
    '../outputs/CSRD_Carbon_Report_2023.xlsx'
)

# ── Weryfikacja ───────────────────────────────────────────────
wb_check = openpyxl.load_workbook('../outputs/CSRD_Carbon_Report_2023.xlsx')
print("Zakładki:", wb_check.sheetnames)
for sheet in wb_check.sheetnames:
    ws = wb_check[sheet]
    print(f"  {sheet}: {ws.max_row} wierszy × {ws.max_column} kolumn")
